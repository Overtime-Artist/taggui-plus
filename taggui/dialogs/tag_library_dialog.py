import csv
from typing import TYPE_CHECKING
from uuid import uuid4

import openpyxl
from openpyxl.styles import Font
from PySide6.QtCore import (QAbstractTableModel, QModelIndex, QPoint, Qt,
                            QTimer, Signal, Slot)
from PySide6.QtGui import QAction, QColor, QKeySequence
from PySide6.QtWidgets import (QAbstractItemView, QApplication, QComboBox,
                                QDialog, QFileDialog,
                                QHeaderView, QHBoxLayout, QLabel, QLineEdit,
                                QListWidget, QListWidgetItem, QMenu, QMessageBox,
                                QTableWidget, QTableWidgetItem,
                                QPushButton, QStyledItemDelegate, QTabWidget,
                                QTableView, QVBoxLayout, QWidget)

from models.tag_library_model import TagLibraryModel
from utils.elided_tooltip import (ElidedToolTipComboBox,
                                  ElidedToolTipListWidget)
from utils.settings import get_tag_separator, get_settings, DEFAULT_SETTINGS
from utils.utils import get_confirmation_dialog_reply, list_with_and, pluralize
from utils.tag_filter import TagFilterLineEdit
from widgets.tag_library_categories_editor import CategoryDialog
from widgets.tag_library_editor import TagLibraryFilterProxyModel

if TYPE_CHECKING:
    from models.image_list_model import ImageListModel


class TagLibraryTableModel(QAbstractTableModel):
    COLUMN_TAG = 0
    COLUMN_CATEGORY = 1

    def __init__(self, tag_library_model: TagLibraryModel):
        super().__init__()
        self._model = tag_library_model
        tag_library_model.modelReset.connect(self._on_model_reset)
        tag_library_model.rowsInserted.connect(self._on_rows_inserted)
        tag_library_model.rowsRemoved.connect(self._on_rows_removed)
        tag_library_model.dataChanged.connect(self._on_data_changed)
        tag_library_model.categories_changed.connect(self._on_categories_changed)

    def rowCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else self._model.rowCount()

    def columnCount(self, parent=QModelIndex()):
        return 0 if parent.isValid() else 2

    def headerData(self, section, orientation,
                   role=Qt.ItemDataRole.DisplayRole):
        if (orientation == Qt.Orientation.Horizontal
                and role == Qt.ItemDataRole.DisplayRole
                and section in (self.COLUMN_TAG, self.COLUMN_CATEGORY)):
            return ('Tag', 'Category')[section]
        return None

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        row, col = index.row(), index.column()
        if row >= self._model.rowCount():
            return None
        source_index = self._model.index(row)
        tag = self._model.data(source_index, Qt.ItemDataRole.EditRole)

        if col == self.COLUMN_TAG:
            if role in (Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.EditRole):
                return tag
            if role == Qt.ItemDataRole.ForegroundRole:
                return self._model.data(source_index, Qt.ItemDataRole.ForegroundRole)

        if col == self.COLUMN_CATEGORY:
            cat = self._model.get_category_for_tag(str(tag)) if tag else None
            if role == Qt.ItemDataRole.DisplayRole:
                return cat['name'] if cat else ''
            if role == Qt.ItemDataRole.EditRole:
                return cat['id'] if cat else ''
            if role == Qt.ItemDataRole.ForegroundRole:
                if cat and cat.get('color'):
                    color = QColor(cat['color'])
                    if color.isValid():
                        return color
        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if (not index.isValid()
                or index.column() != self.COLUMN_CATEGORY
                or role != Qt.ItemDataRole.EditRole):
            return False
        source_index = self._model.index(index.row())
        tag = self._model.data(source_index, Qt.ItemDataRole.EditRole)
        if not tag:
            return False
        if value:
            self._model.assign_category([str(tag)], value)
        else:
            self._model.clear_category([str(tag)])
        return True

    def flags(self, index):
        if not index.isValid():
            return Qt.ItemFlag.NoItemFlags
        base = Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable
        if index.column() == self.COLUMN_CATEGORY:
            base |= Qt.ItemFlag.ItemIsEditable
        return base

    @Slot()
    def _on_model_reset(self):
        self.beginResetModel()
        self.endResetModel()

    @Slot(QModelIndex, int, int)
    def _on_rows_inserted(self, parent, first, last):
        self.beginInsertRows(QModelIndex(), first, last)
        self.endInsertRows()

    @Slot(QModelIndex, int, int)
    def _on_rows_removed(self, parent, first, last):
        self.beginRemoveRows(QModelIndex(), first, last)
        self.endRemoveRows()

    @Slot(QModelIndex, QModelIndex, list)
    def _on_data_changed(self, top_left, bottom_right, roles):
        new_top = self.index(top_left.row(), 0)
        new_bottom = self.index(bottom_right.row(), 1)
        self.dataChanged.emit(new_top, new_bottom, roles)

    @Slot()
    def _on_categories_changed(self):
        row_count = self._model.rowCount()
        if row_count > 0:
            top = self.index(0, self.COLUMN_CATEGORY)
            bottom = self.index(row_count - 1, self.COLUMN_CATEGORY)
            self.dataChanged.emit(
                top, bottom,
                [Qt.ItemDataRole.DisplayRole, Qt.ItemDataRole.ForegroundRole])


class CategoryDelegate(QStyledItemDelegate):
    def __init__(self, tag_library_model: TagLibraryModel, parent=None):
        super().__init__(parent)
        self._model = tag_library_model

    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItem('(none)', '')
        for cat in self._model.get_categories():
            combo.addItem(cat['name'], cat['id'])
        QTimer.singleShot(0, combo.showPopup)
        return combo

    def setEditorData(self, editor, index):
        current_id = index.data(Qt.ItemDataRole.EditRole) or ''
        for i in range(editor.count()):
            if editor.itemData(i) == current_id:
                editor.setCurrentIndex(i)
                return
        editor.setCurrentIndex(0)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentData(), Qt.ItemDataRole.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class TagLibraryTableView(QTableView):
    danbooru_wiki_requested = Signal(str)
    gelbooru_wiki_requested = Signal(str)
    rename_tag_requested = Signal(str)
    assign_category_requested = Signal(str)  # emits the chosen category id
    clear_category_requested = Signal()

    def __init__(self, remove_callback, tag_library_model: TagLibraryModel,
                 selected_tags_callback):
        super().__init__()
        self._remove_callback = remove_callback
        self._tag_library_model = tag_library_model
        self._selected_tags_callback = selected_tags_callback
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_context_menu)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
            self._remove_callback()
            return
        if event.matches(QKeySequence.StandardKey.Copy):
            if self.copy_selected_tags_to_clipboard():
                return
        super().keyPressEvent(event)

    def copy_selected_tags_to_clipboard(self) -> bool:
        """Copy the selected library tags to the clipboard as separator-joined
        text. Returns True if at least one tag was copied."""
        selected_tags = self._selected_tags_callback()
        if not selected_tags:
            return False
        QApplication.clipboard().setText(
            get_tag_separator().join(selected_tags))
        return True

    @Slot(QPoint)
    def _show_context_menu(self, position: QPoint):
        index = self.indexAt(position)
        if not index.isValid():
            return
        # If the right-clicked row is not part of the current selection, select
        # only that row so the menu acts on what the user clicked.
        selected_rows = {selected.row()
                         for selected in self.selectedIndexes()}
        if index.row() not in selected_rows:
            self.selectRow(index.row())
        tag_index = self.model().index(index.row(), TagLibraryTableModel.COLUMN_TAG)
        tag = tag_index.data(Qt.ItemDataRole.DisplayRole)
        if not tag:
            return
        menu = QMenu(self)
        selected_tags = self._selected_tags_callback()
        copy_action = menu.addAction(
            'Copy Tags' if len(selected_tags) > 1 else 'Copy Tag')
        copy_action.setEnabled(bool(selected_tags))
        menu.addSeparator()
        danbooru_action = menu.addAction('View Danbooru Wiki')
        gelbooru_action = menu.addAction('View Gelbooru Wiki')
        menu.addSeparator()
        rename_action = menu.addAction('Rename Tag...')

        categories = self._tag_library_model.get_categories()
        assign_menu = menu.addMenu('Assign Category')
        category_actions = {}
        for category in categories:
            action = assign_menu.addAction(category['name'])
            category_actions[action] = category['id']
        assign_menu.setEnabled(bool(categories))

        any_has_category = any(
            self._tag_library_model.get_category_for_tag(selected_tag)
            is not None
            for selected_tag in selected_tags)
        clear_category_action = menu.addAction('Clear Category')
        clear_category_action.setEnabled(any_has_category)

        menu.addSeparator()
        remove_action = menu.addAction('Remove Selected Tags')
        chosen = menu.exec(self.viewport().mapToGlobal(position))
        if chosen == copy_action:
            self.copy_selected_tags_to_clipboard()
        elif chosen == danbooru_action:
            self.danbooru_wiki_requested.emit(str(tag))
        elif chosen == gelbooru_action:
            self.gelbooru_wiki_requested.emit(str(tag))
        elif chosen == rename_action:
            self.rename_tag_requested.emit(str(tag))
        elif chosen == clear_category_action:
            self.clear_category_requested.emit()
        elif chosen in category_actions:
            self.assign_category_requested.emit(category_actions[chosen])
        elif chosen == remove_action:
            self._remove_callback()


class TagLibrarySortFilterProxyModel(TagLibraryFilterProxyModel):
    def __init__(self, tag_library_model: TagLibraryModel):
        super().__init__()
        self._tag_model = tag_library_model
        # The source model here is the table model, not the tag library model,
        # so category lookups for `category:` filters must use `_tag_model`.
        self.category_lookup_model = tag_library_model
        # When True, rows are ordered by their position in the underlying tag
        # library list instead of alphabetically. New tags are prepended to
        # that list (see TagLibraryModel.add_tags), so a lower source row means
        # a more recently added tag. This is what powers the "Recently added"
        # and "Oldest first" sort modes.
        self.recency_sort = False

    def lessThan(self, left: QModelIndex, right: QModelIndex) -> bool:
        if self.recency_sort:
            # Compare by source-model row so the natural (insertion) order of
            # the tag library list is used. The sort order (ascending vs
            # descending) passed to sort() then decides newest-first vs
            # oldest-first.
            return left.row() < right.row()
        if left.column() == TagLibraryTableModel.COLUMN_CATEGORY:
            left_cat = self._category_sort_key(left)
            right_cat = self._category_sort_key(right)
            if left_cat != right_cat:
                return left_cat < right_cat
            left_tag = (self.sourceModel()
                        .index(left.row(), TagLibraryTableModel.COLUMN_TAG)
                        .data(Qt.ItemDataRole.EditRole) or '')
            right_tag = (self.sourceModel()
                         .index(right.row(), TagLibraryTableModel.COLUMN_TAG)
                         .data(Qt.ItemDataRole.EditRole) or '')
            return left_tag.casefold() < right_tag.casefold()

        left_tag = left.data(Qt.ItemDataRole.EditRole) or ''
        right_tag = right.data(Qt.ItemDataRole.EditRole) or ''
        return left_tag.casefold() < right_tag.casefold()

    def _category_sort_key(self, index: QModelIndex) -> str:
        tag_index = self.sourceModel().index(
            index.row(), TagLibraryTableModel.COLUMN_TAG)
        tag = tag_index.data(Qt.ItemDataRole.EditRole) or ''
        cat = self._tag_model.get_category_for_tag(tag)
        return cat['name'].casefold() if cat else '\xff\xff\xff'


class TagLibraryDialog(QDialog):
    danbooru_wiki_requested = Signal(str)
    gelbooru_wiki_requested = Signal(str)

    def __init__(self, tag_library_model: TagLibraryModel,
                 image_list_model: 'ImageListModel | None' = None, parent=None):
        super().__init__(parent, Qt.WindowType.Window)
        self.tag_library_model = tag_library_model
        self.image_list_model = image_list_model
        self.setWindowTitle('Tag Library')
        self.setMinimumSize(680, 620)
        self.resize(820, 760)

        self.table_model = TagLibraryTableModel(tag_library_model)
        self.proxy_model = TagLibrarySortFilterProxyModel(tag_library_model)
        self.proxy_model.setSourceModel(self.table_model)
        self.proxy_model.setDynamicSortFilter(True)

        tabs = QTabWidget()
        tabs.addTab(self._build_tags_tab(), 'Tags')
        tabs.addTab(self._build_categories_tab(), 'Categories')
        tabs.addTab(self._build_aliases_tab(), 'Aliases')
        tabs.addTab(self._build_implications_tab(), 'Implications')
        tabs.addTab(self._build_profiles_tab(), 'Profiles')

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.addWidget(tabs)

        self.tag_library_model.modelReset.connect(self._update_tag_count)
        self.proxy_model.modelReset.connect(self._update_tag_count)
        self.proxy_model.rowsInserted.connect(self._update_tag_count)
        self.proxy_model.rowsRemoved.connect(self._update_tag_count)
        self.tag_library_model.categories_changed.connect(
            self._refresh_categories_tab)
        self.tag_library_model.categories_changed.connect(
            self._refresh_tags_category_combo_box)
        self.tag_library_model.aliases_changed.connect(
            self._refresh_aliases_tab)
        self.tag_library_model.implications_changed.connect(
            self._refresh_implications_tab)
        self.tag_library_model.profiles_changed.connect(
            self._refresh_profiles_tab)
        self._update_tag_count()
        self._refresh_categories_tab()
        self._refresh_aliases_tab()
        self._refresh_implications_tab()
        self._refresh_profiles_tab()

        # Window-level actions so the wiki keyboard shortcuts also work while
        # this dialog (a separate top-level window) is focused. Their shortcuts
        # are kept in sync with the configured shortcuts via set_wiki_shortcuts.
        self.danbooru_wiki_action = QAction('View Danbooru Wiki', self)
        self.danbooru_wiki_action.setShortcutContext(
            Qt.ShortcutContext.WindowShortcut)
        self.danbooru_wiki_action.triggered.connect(self._request_danbooru_wiki)
        self.addAction(self.danbooru_wiki_action)
        self.gelbooru_wiki_action = QAction('View Gelbooru Wiki', self)
        self.gelbooru_wiki_action.setShortcutContext(
            Qt.ShortcutContext.WindowShortcut)
        self.gelbooru_wiki_action.triggered.connect(self._request_gelbooru_wiki)
        self.addAction(self.gelbooru_wiki_action)

    def set_wiki_shortcuts(self, danbooru_shortcut: QKeySequence,
                           gelbooru_shortcut: QKeySequence):
        """Set the keyboard shortcuts for the wiki actions. Called by the main
        window so these honor any custom shortcuts configured by the user."""
        self.danbooru_wiki_action.setShortcut(danbooru_shortcut)
        self.gelbooru_wiki_action.setShortcut(gelbooru_shortcut)

    def _selected_tag_for_wiki(self) -> str:
        """Return the tag the wiki keyboard shortcut should auto-search.

        If the "Add tag to library..." box is focused, use the text typed in
        it; when that box is empty, fall back to the single highlighted tag in
        the tags table. Otherwise use the single highlighted tag, or '' to open
        the wiki dialog without a preset tag (matching the main window)."""
        if self.add_tag_input.hasFocus():
            text = self.add_tag_input.text().strip()
            if text:
                return text
        selected_tags = self._get_selected_tags()
        return selected_tags[0] if len(selected_tags) == 1 else ''

    @Slot()
    def _request_danbooru_wiki(self):
        self.danbooru_wiki_requested.emit(self._selected_tag_for_wiki())

    @Slot()
    def _request_gelbooru_wiki(self):
        self.gelbooru_wiki_requested.emit(self._selected_tag_for_wiki())

    # ── Tags tab ──────────────────────────────────────────────────────────

    def _build_tags_tab(self) -> QWidget:
        tab = QWidget()

        # Add tag row
        self.add_tag_input = QLineEdit()
        self.add_tag_input.setPlaceholderText('Add Tag')
        self.add_tag_input.setTextMargins(8, 0, 8, 0)
        self.add_tag_input.setClearButtonEnabled(True)
        self.add_tag_button = QPushButton('Add')
        self.add_tag_button.setAutoDefault(False)
        self.add_tag_button.setFixedWidth(50)

        add_row = QHBoxLayout()
        add_row.addWidget(self.add_tag_input)
        add_row.addWidget(self.add_tag_button)

        # Search row
        self.search_input = TagFilterLineEdit('Search Tags')

        # Sort row
        self.sort_combo_box = QComboBox()
        # Each entry stores a (kind, order) tuple as item data. `kind` is one of
        # 'tag', 'category' or 'recent'; `order` is the Qt sort order to apply.
        self.sort_combo_box.addItem(
            'Name (A\u2013Z)',
            ('tag', Qt.SortOrder.AscendingOrder))
        self.sort_combo_box.addItem(
            'Name (Z\u2013A)',
            ('tag', Qt.SortOrder.DescendingOrder))
        self.sort_combo_box.addItem(
            'Category',
            ('category', Qt.SortOrder.AscendingOrder))
        self.sort_combo_box.addItem(
            'Recently added',
            ('recent', Qt.SortOrder.AscendingOrder))
        self.sort_combo_box.addItem(
            'Oldest first',
            ('recent', Qt.SortOrder.DescendingOrder))

        sort_row = QHBoxLayout()
        sort_row.addWidget(QLabel('Sort by'))
        sort_row.addWidget(self.sort_combo_box, stretch=1)

        # Import/export row
        self.import_button = QPushButton('Import...')
        self.import_button.setAutoDefault(False)
        self.export_button = QPushButton('Export...')
        self.export_button.setAutoDefault(False)
        self.template_button = QPushButton('Template...')
        self.template_button.setAutoDefault(False)

        action_row = QHBoxLayout()
        action_row.addStretch()
        action_row.addWidget(self.template_button)
        action_row.addWidget(self.import_button)
        action_row.addWidget(self.export_button)

        self.tags_table = TagLibraryTableView(
            self.request_remove_selected_tags, self.tag_library_model,
            self._get_selected_tags)
        self.tags_table.setModel(self.proxy_model)
        self.tags_table.setSortingEnabled(True)
        self.tags_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.tags_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tags_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tags_table.setAlternatingRowColors(True)
        self.tags_table.setWordWrap(False)

        header = self.tags_table.horizontalHeader()
        header.setSectionResizeMode(TagLibraryTableModel.COLUMN_TAG,
                                    QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(TagLibraryTableModel.COLUMN_CATEGORY,
                                    QHeaderView.ResizeMode.Fixed)
        header.resizeSection(TagLibraryTableModel.COLUMN_CATEGORY, 160)
        header.setSortIndicatorShown(True)
        # Guard so programmatic sorts (driven by the sort dropdown below) don't
        # get treated as if the user clicked a column header.
        self._applying_sort_mode = False

        self.tags_table.verticalHeader().setVisible(False)

        self.category_delegate = CategoryDelegate(self.tag_library_model,
                                                  self.tags_table)
        self.tags_table.setItemDelegateForColumn(
            TagLibraryTableModel.COLUMN_CATEGORY, self.category_delegate)

        # Count label
        self.tag_count_label = QLabel()

        self.remove_selected_button = QPushButton('Remove Selected Tags')
        self.remove_selected_button.setAutoDefault(False)

        # Category assignment controls
        self.tags_category_combo_box = ElidedToolTipComboBox()
        self.tags_category_combo_box.setSizeAdjustPolicy(
            ElidedToolTipComboBox.SizeAdjustPolicy
            .AdjustToMinimumContentsLengthWithIcon)
        self.assign_category_button = QPushButton('Assign Category')
        self.assign_category_button.setAutoDefault(False)
        self.clear_category_button = QPushButton('Clear Category')
        self.clear_category_button.setAutoDefault(False)

        category_row = QHBoxLayout()
        category_row.addWidget(QLabel('Selected category'))
        category_row.addWidget(self.tags_category_combo_box, stretch=1)
        category_row.addWidget(self.assign_category_button)
        category_row.addWidget(self.clear_category_button)

        layout = QVBoxLayout(tab)
        layout.addLayout(add_row)
        layout.addWidget(self.search_input)
        layout.addLayout(sort_row)
        layout.addLayout(action_row)
        layout.addWidget(self.tags_table)
        layout.addWidget(self.tag_count_label)
        layout.addLayout(category_row)
        layout.addWidget(self.remove_selected_button)

        # Connections
        self.add_tag_input.returnPressed.connect(self._add_tag)
        self.add_tag_button.clicked.connect(self._add_tag)
        self.search_input.textChanged.connect(self._apply_tag_filter)
        self.search_input.textChanged.connect(self._update_tag_count)
        self.template_button.clicked.connect(self._download_template)
        self.import_button.clicked.connect(self._import_library)
        self.export_button.clicked.connect(self._export_library)
        self.remove_selected_button.clicked.connect(
            self.request_remove_selected_tags)
        self.assign_category_button.clicked.connect(
            self._request_assign_category_to_selected_tags)
        self.clear_category_button.clicked.connect(
            self._request_clear_category_for_selected_tags)
        self.tags_table.clicked.connect(self._on_table_clicked)
        self.tags_table.danbooru_wiki_requested.connect(
            self.danbooru_wiki_requested.emit)
        self.tags_table.gelbooru_wiki_requested.connect(
            self.gelbooru_wiki_requested.emit)
        self.tags_table.rename_tag_requested.connect(self._rename_tag)
        self.tags_table.assign_category_requested.connect(
            self._assign_category_to_selected_tags)
        self.tags_table.clear_category_requested.connect(
            self._request_clear_category_for_selected_tags)
        self.sort_combo_box.currentIndexChanged.connect(self._apply_sort_mode)
        self.tags_table.horizontalHeader().sortIndicatorChanged.connect(
            self._on_sort_indicator_changed)

        self._refresh_tags_category_combo_box()
        # Restore the user's last-used sort choice (persisted across restarts),
        # then apply it so the dropdown and the table start in agreement.
        self._restore_saved_sort_mode()
        self._apply_sort_mode()

        return tab

    def _restore_saved_sort_mode(self):
        """Select the sort option saved from a previous session, if present."""
        settings = get_settings()
        saved_label = settings.value(
            'tag_library_sort_by',
            defaultValue=DEFAULT_SETTINGS['tag_library_sort_by'], type=str)
        index = self.sort_combo_box.findText(saved_label)
        if index != -1:
            self.sort_combo_box.blockSignals(True)
            self.sort_combo_box.setCurrentIndex(index)
            self.sort_combo_box.blockSignals(False)

    @Slot()
    def _apply_sort_mode(self):
        """Sort the tags table according to the current sort dropdown choice."""
        data = self.sort_combo_box.currentData()
        if not data:
            return
        # Persist the choice so it is restored on the next application launch.
        get_settings().setValue('tag_library_sort_by',
                                self.sort_combo_box.currentText())
        kind, order = data
        self._applying_sort_mode = True
        try:
            self.proxy_model.recency_sort = (kind == 'recent')
            if kind == 'category':
                column = TagLibraryTableModel.COLUMN_CATEGORY
            else:
                column = TagLibraryTableModel.COLUMN_TAG
            self.tags_table.sortByColumn(column, order)
            # Force a re-sort even if the column/order happen to match the
            # previous ones (e.g. switching between "Name (A–Z)" and "Recently
            # added", which both use column 0 ascending): only the recency flag
            # changed, so lessThan() must run again.
            self.proxy_model.invalidate()
        finally:
            self._applying_sort_mode = False

    @Slot(int, Qt.SortOrder)
    def _on_sort_indicator_changed(self, column: int, order: Qt.SortOrder):
        """Keep the sort dropdown in sync when the user clicks a column header."""
        if self._applying_sort_mode:
            return
        was_recency = self.proxy_model.recency_sort
        # A header click always means a plain column sort, never recency.
        self.proxy_model.recency_sort = False
        if was_recency:
            # The click already sorted using the (now stale) recency comparison,
            # so re-sort by the clicked column.
            self.proxy_model.invalidate()
        target_kind = ('category'
                       if column == TagLibraryTableModel.COLUMN_CATEGORY
                       else 'tag')
        self.sort_combo_box.blockSignals(True)
        for index in range(self.sort_combo_box.count()):
            data = self.sort_combo_box.itemData(index)
            if not data:
                continue
            kind, item_order = data
            if kind == target_kind and (target_kind == 'category'
                                        or item_order == order):
                self.sort_combo_box.setCurrentIndex(index)
                break
        self.sort_combo_box.blockSignals(False)
        # Signals were blocked above, so persist the choice explicitly here.
        get_settings().setValue('tag_library_sort_by',
                                self.sort_combo_box.currentText())


    # ── Categories tab ────────────────────────────────────────────────────

    def _build_categories_tab(self) -> QWidget:
        tab = QWidget()

        self.categories_list = ElidedToolTipListWidget()
        self.categories_list.setDragDropMode(
            QAbstractItemView.DragDropMode.InternalMove)
        self.categories_list.setDefaultDropAction(Qt.DropAction.MoveAction)
        self.categories_list.setSelectionMode(
            QAbstractItemView.SelectionMode.SingleSelection)

        self.category_count_label = QLabel()

        self.add_category_button = QPushButton('Add Category')
        self.add_category_button.setAutoDefault(False)
        self.edit_category_button = QPushButton('Edit Category')
        self.edit_category_button.setAutoDefault(False)
        self.remove_category_button = QPushButton('Remove Category')
        self.remove_category_button.setAutoDefault(False)

        cat_buttons_row = QHBoxLayout()
        cat_buttons_row.addWidget(self.edit_category_button, stretch=1)
        cat_buttons_row.addWidget(self.remove_category_button, stretch=1)

        layout = QVBoxLayout(tab)
        layout.addWidget(self.categories_list)
        layout.addWidget(self.category_count_label)
        layout.addWidget(self.add_category_button)
        layout.addLayout(cat_buttons_row)

        self.categories_list.model().rowsMoved.connect(
            self._persist_category_order)
        self.add_category_button.clicked.connect(self._request_add_category)
        self.edit_category_button.clicked.connect(self._request_edit_category)
        self.remove_category_button.clicked.connect(
            self._request_remove_category)

        return tab

    def _build_aliases_tab(self) -> QWidget:
        tab = QWidget()

        self.new_alias_input = QLineEdit()
        self.new_alias_input.setPlaceholderText('Aliases')
        self.new_alias_input.setTextMargins(8, 0, 8, 0)
        arrow_label = QLabel('→')
        arrow_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.new_canonical_input = QLineEdit()
        self.new_canonical_input.setPlaceholderText('Canonical Tag')
        self.new_canonical_input.setTextMargins(8, 0, 8, 0)
        self.add_alias_button = QPushButton('Add')
        self.add_alias_button.setAutoDefault(False)
        self.add_alias_button.setFixedWidth(50)

        add_row = QHBoxLayout()
        add_row.addWidget(self.new_alias_input, stretch=2)
        add_row.addWidget(arrow_label)
        add_row.addWidget(self.new_canonical_input, stretch=2)
        add_row.addWidget(self.add_alias_button)

        self.alias_search_input = QLineEdit()
        self.alias_search_input.setPlaceholderText('Search Aliases')
        self.alias_search_input.setTextMargins(8, 0, 8, 0)
        self.alias_search_input.setClearButtonEnabled(True)

        self.aliases_table = QTableWidget(0, 2)
        self.aliases_table.setHorizontalHeaderLabels(['Alias', 'Canonical Tag'])
        self.aliases_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.aliases_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.aliases_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.aliases_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.aliases_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.aliases_table.verticalHeader().setVisible(False)
        self.aliases_table.setSortingEnabled(True)

        self.alias_count_label = QLabel()

        self.edit_alias_button = QPushButton('Edit Selected')
        self.edit_alias_button.setAutoDefault(False)
        self.remove_alias_button = QPushButton('Remove Selected Aliases')
        self.remove_alias_button.setAutoDefault(False)

        buttons_row = QHBoxLayout()
        buttons_row.addWidget(self.edit_alias_button)
        buttons_row.addWidget(self.remove_alias_button)

        layout = QVBoxLayout(tab)
        layout.addLayout(add_row)
        layout.addWidget(self.alias_search_input)
        layout.addWidget(self.aliases_table)
        layout.addWidget(self.alias_count_label)
        layout.addLayout(buttons_row)

        self.add_alias_button.clicked.connect(self._add_alias)
        self.new_alias_input.returnPressed.connect(self._add_alias)
        self.new_canonical_input.returnPressed.connect(self._add_alias)
        self.alias_search_input.textChanged.connect(self._filter_aliases_table)
        self.edit_alias_button.clicked.connect(self._edit_selected_alias)
        self.remove_alias_button.clicked.connect(self._remove_selected_aliases)

        return tab

    def _build_implications_tab(self) -> QWidget:
        tab = QWidget()

        self.implication_tag_input = QLineEdit()
        self.implication_tag_input.setPlaceholderText('Tags')
        self.implication_tag_input.setTextMargins(8, 0, 8, 0)
        implies_label = QLabel('implies')
        implies_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.implication_implied_input = QLineEdit()
        self.implication_implied_input.setPlaceholderText('Implied Tags')
        self.implication_implied_input.setTextMargins(8, 0, 8, 0)
        self.add_implication_button = QPushButton('Add')
        self.add_implication_button.setAutoDefault(False)
        self.add_implication_button.setFixedWidth(50)

        add_row = QHBoxLayout()
        add_row.addWidget(self.implication_tag_input, stretch=2)
        add_row.addWidget(implies_label)
        add_row.addWidget(self.implication_implied_input, stretch=1)
        add_row.addWidget(self.add_implication_button)

        self.implication_search_input = QLineEdit()
        self.implication_search_input.setPlaceholderText(
            'Search Implications')
        self.implication_search_input.setTextMargins(8, 0, 8, 0)
        self.implication_search_input.setClearButtonEnabled(True)

        self.implications_table = QTableWidget(0, 2)
        self.implications_table.setHorizontalHeaderLabels(['Tag', 'Implies'])
        self.implications_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.implications_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.implications_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.implications_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.implications_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.implications_table.verticalHeader().setVisible(False)
        self.implications_table.setSortingEnabled(True)

        self.implication_count_label = QLabel()
        self.edit_implication_button = QPushButton('Edit Selected')
        self.edit_implication_button.setAutoDefault(False)
        self.remove_implication_button = QPushButton('Remove Selected Rules')
        self.remove_implication_button.setAutoDefault(False)

        buttons_row = QHBoxLayout()
        buttons_row.addWidget(self.edit_implication_button)
        buttons_row.addWidget(self.remove_implication_button)

        layout = QVBoxLayout(tab)
        layout.addLayout(add_row)
        layout.addWidget(self.implication_search_input)
        layout.addWidget(self.implications_table)
        layout.addWidget(self.implication_count_label)
        layout.addLayout(buttons_row)

        self.add_implication_button.clicked.connect(self._add_implication)
        self.implication_tag_input.returnPressed.connect(self._add_implication)
        self.implication_implied_input.returnPressed.connect(self._add_implication)
        self.implication_search_input.textChanged.connect(
            self._filter_implications_table)
        self.edit_implication_button.clicked.connect(
            self._edit_selected_implication)
        self.remove_implication_button.clicked.connect(
            self._remove_selected_implication_rules)

        return tab

    def _build_profiles_tab(self) -> QWidget:
        tab = QWidget()

        self.profiles_list = QListWidget()
        self.profile_count_label = QLabel()

        self.add_profile_button = QPushButton('Add Profile')
        self.add_profile_button.setAutoDefault(False)
        self.rename_profile_button = QPushButton('Rename Profile')
        self.rename_profile_button.setAutoDefault(False)
        self.remove_profile_button = QPushButton('Remove Profile')
        self.remove_profile_button.setAutoDefault(False)
        self.apply_profile_button = QPushButton('Apply Profile to Directory...')
        self.apply_profile_button.setAutoDefault(False)
        self.revert_profile_button = QPushButton('Revert Profile from Directory...')
        self.revert_profile_button.setAutoDefault(False)

        left_layout = QVBoxLayout()
        left_layout.addWidget(self.profiles_list)
        left_layout.addWidget(self.profile_count_label)
        left_layout.addWidget(self.add_profile_button)
        left_layout.addWidget(self.rename_profile_button)
        left_layout.addWidget(self.remove_profile_button)
        left_layout.addWidget(self.apply_profile_button)
        left_layout.addWidget(self.revert_profile_button)
        left_layout.addStretch()

        left_panel = QWidget()
        left_panel.setLayout(left_layout)

        self.profile_right_panel = QWidget()
        right_layout = QVBoxLayout(self.profile_right_panel)

        self.profile_mappings_label = QLabel('Mappings')

        self.profile_original_input = QLineEdit()
        self.profile_original_input.setPlaceholderText('Original Tag')
        self.profile_original_input.setTextMargins(8, 0, 8, 0)
        profile_arrow_label = QLabel('→')
        profile_arrow_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.profile_replacement_input = QLineEdit()
        self.profile_replacement_input.setPlaceholderText('Replacement')
        self.profile_replacement_input.setTextMargins(8, 0, 8, 0)
        self.add_mapping_button = QPushButton('Add')
        self.add_mapping_button.setAutoDefault(False)
        self.add_mapping_button.setFixedWidth(50)

        add_row = QHBoxLayout()
        add_row.addWidget(self.profile_original_input, stretch=2)
        add_row.addWidget(profile_arrow_label)
        add_row.addWidget(self.profile_replacement_input, stretch=2)
        add_row.addWidget(self.add_mapping_button)

        self.profile_mapping_search_input = QLineEdit()
        self.profile_mapping_search_input.setPlaceholderText('Search Mappings')
        self.profile_mapping_search_input.setTextMargins(8, 0, 8, 0)
        self.profile_mapping_search_input.setClearButtonEnabled(True)

        self.profile_mappings_table = QTableWidget(0, 2)
        self.profile_mappings_table.setHorizontalHeaderLabels(
            ['Original Tag', 'Replacement'])
        self.profile_mappings_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Stretch)
        self.profile_mappings_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch)
        self.profile_mappings_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.profile_mappings_table.setSelectionMode(
            QAbstractItemView.SelectionMode.ExtendedSelection)
        self.profile_mappings_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.profile_mappings_table.verticalHeader().setVisible(False)
        self.profile_mappings_table.setSortingEnabled(True)

        self.profile_mapping_count_label = QLabel('0 Mappings')

        self.edit_mapping_button = QPushButton('Edit Selected')
        self.edit_mapping_button.setAutoDefault(False)
        self.remove_mapping_button = QPushButton('Remove Selected Mappings')
        self.remove_mapping_button.setAutoDefault(False)

        buttons_row = QHBoxLayout()
        buttons_row.addWidget(self.edit_mapping_button)
        buttons_row.addWidget(self.remove_mapping_button)

        right_layout.addWidget(self.profile_mappings_label)
        right_layout.addLayout(add_row)
        right_layout.addWidget(self.profile_mapping_search_input)
        right_layout.addWidget(self.profile_mappings_table)
        right_layout.addWidget(self.profile_mapping_count_label)
        right_layout.addLayout(buttons_row)

        layout = QHBoxLayout(tab)
        layout.addWidget(left_panel, stretch=1)
        layout.addWidget(self.profile_right_panel, stretch=1)

        self.profiles_list.currentItemChanged.connect(self._on_profile_selected)
        self.add_profile_button.clicked.connect(self._add_profile)
        self.rename_profile_button.clicked.connect(self._rename_profile)
        self.remove_profile_button.clicked.connect(self._remove_profile)
        self.apply_profile_button.clicked.connect(self._apply_profile)
        self.revert_profile_button.clicked.connect(self._revert_profile)
        self.add_mapping_button.clicked.connect(self._add_mapping)
        self.profile_original_input.returnPressed.connect(self._add_mapping)
        self.profile_replacement_input.returnPressed.connect(self._add_mapping)
        self.edit_mapping_button.clicked.connect(self._edit_selected_mapping)
        self.remove_mapping_button.clicked.connect(self._remove_selected_mappings)
        self.profile_mapping_search_input.textChanged.connect(
            self._filter_profile_mappings_table)

        return tab

    # ── Tags tab slots ────────────────────────────────────────────────────

    def _get_selected_profile_name(self) -> str | None:
        item = self.profiles_list.currentItem()
        return item.text() if item else None

    @Slot()
    def _refresh_profiles_tab(self):
        current_name = self._get_selected_profile_name()
        self.profiles_list.blockSignals(True)
        self.profiles_list.clear()
        profiles = self.tag_library_model.get_profiles()
        for name in profiles:
            self.profiles_list.addItem(name)
        count = len(profiles)
        self.profile_count_label.setText(
            f'{count} Profile{"s" if count != 1 else ""}')
        if current_name:
            items = self.profiles_list.findItems(
                current_name, Qt.MatchFlag.MatchExactly)
            if items:
                self.profiles_list.setCurrentItem(items[0])
        self.profiles_list.blockSignals(False)
        self._refresh_profile_mappings()

    @Slot()
    def _refresh_profile_mappings(self):
        profile_name = self._get_selected_profile_name()
        has_profile = profile_name is not None
        self.profile_right_panel.setEnabled(has_profile)
        self.profile_mappings_label.setText(
            f'Mappings for: {profile_name}' if has_profile else 'Mappings')
        self.profile_mappings_table.setSortingEnabled(False)
        self.profile_mappings_table.setRowCount(0)
        if has_profile:
            profiles = self.tag_library_model.get_profiles()
            mapping = profiles.get(profile_name, {})
            for original, replacement in sorted(mapping.items()):
                row = self.profile_mappings_table.rowCount()
                self.profile_mappings_table.insertRow(row)
                self.profile_mappings_table.setItem(
                    row, 0, QTableWidgetItem(original))
                self.profile_mappings_table.setItem(
                    row, 1, QTableWidgetItem(replacement))
            count = len(mapping)
            self.profile_mapping_count_label.setText(
                f'{count} Mapping{"s" if count != 1 else ""}')
        else:
            self.profile_mapping_count_label.setText('0 Mappings')
        self.profile_mappings_table.setSortingEnabled(True)
        self._filter_profile_mappings_table()

    @Slot()
    def _on_profile_selected(self):
        self._refresh_profile_mappings()

    @Slot()
    def _add_profile(self):
        from PySide6.QtWidgets import QInputDialog

        name, ok = QInputDialog.getText(self, 'Add Profile', 'Profile name:')
        if not ok or not name.strip():
            return
        name = name.strip()
        if not self.tag_library_model.add_profile(name):
            QMessageBox.warning(
                self, 'Add Profile',
                f'A profile named "{name}" already exists.')
            return
        items = self.profiles_list.findItems(name, Qt.MatchFlag.MatchExactly)
        if items:
            self.profiles_list.setCurrentItem(items[0])

    @Slot()
    def _rename_profile(self):
        from PySide6.QtWidgets import QInputDialog

        current = self._get_selected_profile_name()
        if not current:
            return
        new_name, ok = QInputDialog.getText(
            self, 'Rename Profile', 'New profile name:', text=current)
        if not ok or not new_name.strip() or new_name.strip() == current:
            return
        new_name = new_name.strip()
        if not self.tag_library_model.rename_profile(current, new_name):
            QMessageBox.warning(
                self, 'Rename Profile',
                f'A profile named "{new_name}" already exists.')

    @Slot()
    def _remove_profile(self):
        name = self._get_selected_profile_name()
        if not name:
            return
        reply = get_confirmation_dialog_reply(
            title='Remove Profile',
            question=f'Remove profile "{name}"? This cannot be undone.')
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.tag_library_model.remove_profile(name)

    @Slot()
    def _add_mapping(self):
        profile_name = self._get_selected_profile_name()
        if not profile_name:
            return
        original = self.profile_original_input.text().strip()
        replacement = self.profile_replacement_input.text().strip()
        if not original:
            return
        self.tag_library_model.set_profile_mapping(
            profile_name, original, replacement)
        self.profile_original_input.clear()
        self.profile_replacement_input.clear()

    @Slot()
    def _edit_selected_mapping(self):
        profile_name = self._get_selected_profile_name()
        if not profile_name:
            return
        selected_rows = sorted(set(
            idx.row() for idx in self.profile_mappings_table.selectedIndexes()))
        if len(selected_rows) != 1:
            return
        row = selected_rows[0]
        orig_item = self.profile_mappings_table.item(row, 0)
        repl_item = self.profile_mappings_table.item(row, 1)
        if not orig_item:
            return
        self.profile_original_input.setText(orig_item.text())
        self.profile_replacement_input.setText(
            repl_item.text() if repl_item else '')
        self.tag_library_model.remove_profile_mappings(
            profile_name, [orig_item.text()])
        self.profile_original_input.setFocus()

    @Slot()
    def _remove_selected_mappings(self):
        profile_name = self._get_selected_profile_name()
        if not profile_name:
            return
        selected_rows = set(
            idx.row() for idx in self.profile_mappings_table.selectedIndexes())
        if not selected_rows:
            return
        originals = []
        for row in selected_rows:
            item = self.profile_mappings_table.item(row, 0)
            if item:
                originals.append(item.text())
        self.tag_library_model.remove_profile_mappings(profile_name, originals)

    @Slot()
    def _filter_profile_mappings_table(self):
        text = self.profile_mapping_search_input.text().lower()
        for row in range(self.profile_mappings_table.rowCount()):
            item0 = self.profile_mappings_table.item(row, 0)
            item1 = self.profile_mappings_table.item(row, 1)
            t0 = item0.text().lower() if item0 else ''
            t1 = item1.text().lower() if item1 else ''
            self.profile_mappings_table.setRowHidden(
                row, text not in t0 and text not in t1)

    @Slot()
    def _apply_profile(self):
        profile_name = self._get_selected_profile_name()
        if not profile_name:
            return
        profiles = self.tag_library_model.get_profiles()
        mapping = profiles.get(profile_name, {})
        if not mapping:
            QMessageBox.information(
                self, 'Apply Profile',
                f'Profile "{profile_name}" has no mappings.')
            return
        if self.image_list_model is None:
            QMessageBox.warning(
                self, 'Apply Profile',
                'No directory is currently loaded.')
            return
        images = self.image_list_model.images
        if not images:
            QMessageBox.warning(
                self, 'Apply Profile',
                'No images are loaded in the current directory.')
            return

        affected = sum(
            1 for image in images
            if any(tag in mapping for tag in image.tags))
        if affected == 0:
            QMessageBox.information(
                self, 'Apply Profile',
                f'No images in the current directory contain tags from profile '
                f'"{profile_name}".')
            return

        reply = get_confirmation_dialog_reply(
            title='Apply Profile',
            question=(f'Apply profile "{profile_name}" to {affected} image(s) '
                      f'in the current directory?\n\n'
                      f'This will replace {len(mapping)} tag(s) across all '
                      'matching images. This action can be undone.'))
        if reply != QMessageBox.StandardButton.Yes:
            return

        for original, replacement in mapping.items():
            if replacement:
                self.image_list_model.rename_tags([original], replacement)

        QMessageBox.information(
            self, 'Profile Applied',
            f'Profile "{profile_name}" applied successfully.')

    @Slot()
    def _revert_profile(self):
        profile_name = self._get_selected_profile_name()
        if not profile_name:
            return
        profiles = self.tag_library_model.get_profiles()
        mapping = profiles.get(profile_name, {})
        # Only mappings with a non-empty replacement can be reverted
        revertible = {orig: repl for orig, repl in mapping.items() if repl}
        if not revertible:
            QMessageBox.information(
                self, 'Revert Profile',
                f'Profile "{profile_name}" has no mappings that can be reverted.')
            return
        if self.image_list_model is None:
            QMessageBox.warning(
                self, 'Revert Profile',
                'No directory is currently loaded.')
            return
        images = self.image_list_model.images
        if not images:
            QMessageBox.warning(
                self, 'Revert Profile',
                'No images are loaded in the current directory.')
            return

        # Check for ambiguous reversals: if two originals share the same
        # replacement, reverting is undefined — warn and abort.
        replacements_seen = {}
        for orig, repl in revertible.items():
            if repl in replacements_seen:
                QMessageBox.warning(
                    self, 'Revert Profile',
                    f'Cannot revert: replacement tag "{repl}" is used by more '
                    f'than one mapping ("{replacements_seen[repl]}" and '
                    f'"{orig}"). Remove the ambiguity before reverting.')
                return
            replacements_seen[repl] = orig

        # Inverted mapping: replacement → original
        inverted = {repl: orig for orig, repl in revertible.items()}

        affected = sum(
            1 for image in images
            if any(tag in inverted for tag in image.tags))
        if affected == 0:
            QMessageBox.information(
                self, 'Revert Profile',
                f'No images in the current directory contain replacement tags '
                f'from profile "{profile_name}".')
            return

        reply = get_confirmation_dialog_reply(
            title='Revert Profile',
            question=(f'Revert profile "{profile_name}" on {affected} image(s) '
                      f'in the current directory?\n\n'
                      f'This will restore {len(inverted)} tag(s) to their '
                      'original values. This action can be undone.'))
        if reply != QMessageBox.StandardButton.Yes:
            return

        for replacement, original in inverted.items():
            self.image_list_model.rename_tags([replacement], original)

        QMessageBox.information(
            self, 'Profile Reverted',
            f'Profile "{profile_name}" reverted successfully.')

    @Slot(QModelIndex)
    def _on_table_clicked(self, index: QModelIndex):
        if index.column() == TagLibraryTableModel.COLUMN_CATEGORY:
            self.tags_table.edit(index)

    @Slot()
    def _add_tag(self):
        text = self.add_tag_input.text().strip()
        if not text:
            return
        # Add the tag and, if it is genuinely new to the library, immediately
        # show the same category-assignment prompt used elsewhere so the user
        # can categorize it without having to filter for and select it first.
        from widgets.image_tags_editor import add_new_tag_library_tags_with_prompt
        add_new_tag_library_tags_with_prompt(self, self.tag_library_model,
                                             [text])
        self.add_tag_input.clear()

    def _get_selected_tags(self) -> list[str]:
        selected_rows = set(index.row()
                            for index in self.tags_table.selectedIndexes())
        tags = []
        for row in selected_rows:
            tag_index = self.proxy_model.index(row, TagLibraryTableModel.COLUMN_TAG)
            tag = tag_index.data(Qt.ItemDataRole.DisplayRole)
            if tag:
                tags.append(tag)
        tags = list(dict.fromkeys(tags))
        tags.sort(key=str.casefold)
        return tags

    @Slot()
    def _refresh_tags_category_combo_box(self):
        previous_category_id = self.tags_category_combo_box.currentData()
        categories = self.tag_library_model.get_categories()
        self.tags_category_combo_box.blockSignals(True)
        self.tags_category_combo_box.clear()
        self.tags_category_combo_box.addItem('Choose category...', '')
        for category in categories:
            self.tags_category_combo_box.addItem(category['name'],
                                                 category['id'])
        selected_index = 0
        if previous_category_id:
            for index in range(self.tags_category_combo_box.count()):
                if (self.tags_category_combo_box.itemData(index)
                        == previous_category_id):
                    selected_index = index
                    break
        self.tags_category_combo_box.setCurrentIndex(selected_index)
        self.tags_category_combo_box.blockSignals(False)
        self.tags_category_combo_box.update_elided_tooltip()

    @Slot()
    def _request_assign_category_to_selected_tags(self):
        category_id = self.tags_category_combo_box.currentData()
        if not category_id:
            return
        self._assign_category_to_selected_tags(category_id)

    @Slot(str)
    def _assign_category_to_selected_tags(self, category_id: str):
        tags = self._get_selected_tags()
        if not tags or not category_id:
            return
        self.tag_library_model.assign_category(tags, category_id)

    @Slot()
    def _request_clear_category_for_selected_tags(self):
        tags = self._get_selected_tags()
        if not tags:
            return
        self.tag_library_model.clear_category(tags)

    @Slot(str)
    def _rename_tag(self, old_tag: str):
        from PySide6.QtWidgets import QInputDialog

        old_tag = str(old_tag).strip()
        if not old_tag:
            return
        new_tag, ok = QInputDialog.getText(
            self, 'Rename Tag', f'New name for "{old_tag}":', text=old_tag)
        if not ok:
            return
        new_tag = new_tag.strip()
        if not new_tag or new_tag == old_tag:
            return
        # Rename in the local tag library (keeps category / aliases /
        # implications pointed at the new name).
        self.tag_library_model.rename_tags([old_tag], new_tag)
        # Rename everywhere in the currently loaded images so the caption
        # files are updated too.
        if self.image_list_model is not None:
            self.image_list_model.rename_tags([old_tag], new_tag)

    @Slot()
    def request_remove_selected_tags(self):
        tags = self._get_selected_tags()
        if not tags:
            return
        question = f'Remove {pluralize("tag", len(tags))} '
        if len(tags) < 10:
            quoted = [f'"{t}"' for t in tags]
            question += f'{list_with_and(quoted)} from the Tag Library?'
        else:
            question += f'from the Tag Library? ({len(tags)} selected)'
        reply = get_confirmation_dialog_reply(title='Remove from Tag Library',
                                              question=question)
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.tag_library_model.remove_tags(tags)
        self._offer_to_remove_tags_from_loaded_images(tags)

    def _offer_to_remove_tags_from_loaded_images(self, tags: list[str]):
        """After a library removal, offer to also delete the tags from all
        loaded image captions. Otherwise the tags stay on the images and are
        re-added to the library the next time the directory is loaded."""
        if self.image_list_model is None:
            return
        tags_in_images = [
            tag for tag in tags
            if any(tag in image.tags for image in self.image_list_model.images)
        ]
        if not tags_in_images:
            return
        count = len(tags_in_images)
        if count < 10:
            quoted = [f'"{t}"' for t in tags_in_images]
            tag_phrase = list_with_and(quoted)
        else:
            tag_phrase = f'{count} of the removed tags'
        noun = 'tag' if count == 1 else 'tags'
        pronoun = 'it' if count == 1 else 'them'
        verb = 'is' if count == 1 else 'are'
        question = (
            f'The {noun} {tag_phrase} {verb} still present in your loaded '
            f'images. Removing from the library does not delete {pronoun} from '
            f'the images, and {pronoun} will be re-added to the library the '
            f'next time this directory is loaded.\n\n'
            f'Also remove {pronoun} from all loaded images?')
        reply = get_confirmation_dialog_reply(title='Remove from Images',
                                              question=question)
        if reply == QMessageBox.StandardButton.Yes:
            self.image_list_model.delete_tags(tags_in_images)

    @Slot()
    def _apply_tag_filter(self):
        self.proxy_model.set_filter(self.search_input.parse_filter_text())

    @Slot()
    def _update_tag_count(self):
        total = self.tag_library_model.rowCount()
        shown = self.proxy_model.rowCount()
        self.tag_count_label.setText(f'{shown} / {total} Library Tags')

    # ── Import / Export ───────────────────────────────────────────────────

    @Slot()
    def _export_library(self):
        model = self.tag_library_model
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Export Tag Library', 'tag_library.xlsx',
            'Excel Files (*.xlsx);;All Files (*)')
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()

            ws_tags = wb.active
            ws_tags.title = 'Tags'
            ws_tags.append(['tag', 'category'])
            ws_tags['A1'].font = Font(bold=True)
            ws_tags['B1'].font = Font(bold=True)
            cat_by_id = {c['id']: c for c in model.categories}
            for tag in model.tags:
                cat = cat_by_id.get(model.category_by_tag.get(tag, ''))
                ws_tags.append([tag, cat['name'] if cat else ''])

            ws_cats = wb.create_sheet('Categories')
            ws_cats.append(['name', 'color (hex)'])
            ws_cats['A1'].font = Font(bold=True)
            ws_cats['B1'].font = Font(bold=True)
            for cat in model.categories:
                ws_cats.append([cat['name'], cat.get('color', '')])

            ws_aliases = wb.create_sheet('Aliases')
            ws_aliases.append(['alias', 'canonical'])
            ws_aliases['A1'].font = Font(bold=True)
            ws_aliases['B1'].font = Font(bold=True)
            for alias, canonical in sorted(model.aliases.items()):
                ws_aliases.append([alias, canonical])

            ws_impl = wb.create_sheet('Implications')
            ws_impl.append(['tag', 'implies'])
            ws_impl['A1'].font = Font(bold=True)
            ws_impl['B1'].font = Font(bold=True)
            for tag, implied_tags in sorted(model.implications.items()):
                ws_impl.append([tag, ', '.join(implied_tags)])

            ws_profiles = wb.create_sheet('Profiles')
            ws_profiles.append(['profile', 'original', 'replacement'])
            ws_profiles['A1'].font = Font(bold=True)
            ws_profiles['B1'].font = Font(bold=True)
            ws_profiles['C1'].font = Font(bold=True)
            for profile_name, mapping in sorted(model.profiles.items()):
                for original, replacement in sorted(mapping.items()):
                    ws_profiles.append([profile_name, original, replacement])

            wb.save(file_path)
            QMessageBox.information(
                self, 'Export Successful',
                f'Tag Library exported to:\n{file_path}')
        except Exception as e:
            QMessageBox.warning(self, 'Export Failed',
                                f'Failed to export Tag Library:\n{e}')

    @Slot()
    def _download_template(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, 'Save Tag Library Template', 'tag_library_template.xlsx',
            'Excel Files (*.xlsx);;All Files (*)')
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws_tags = wb.active
            ws_tags.title = 'Tags'
            ws_tags.append(['tag', 'category'])
            ws_tags['A1'].font = Font(bold=True)
            ws_tags['B1'].font = Font(bold=True)

            ws_cats = wb.create_sheet('Categories')
            ws_cats.append(['name', 'color (hex)'])
            ws_cats['A1'].font = Font(bold=True)
            ws_cats['B1'].font = Font(bold=True)

            ws_aliases = wb.create_sheet('Aliases')
            ws_aliases.append(['alias', 'canonical'])
            ws_aliases['A1'].font = Font(bold=True)
            ws_aliases['B1'].font = Font(bold=True)

            ws_impl = wb.create_sheet('Implications')
            ws_impl.append(['tag', 'implies'])
            ws_impl['A1'].font = Font(bold=True)
            ws_impl['B1'].font = Font(bold=True)

            ws_profiles = wb.create_sheet('Profiles')
            ws_profiles.append(['profile', 'original', 'replacement'])
            ws_profiles['A1'].font = Font(bold=True)
            ws_profiles['B1'].font = Font(bold=True)
            ws_profiles['C1'].font = Font(bold=True)

            wb.save(file_path)
            QMessageBox.information(
                self, 'Template Saved',
                f'Template saved to:\n{file_path}')
        except Exception as e:
            QMessageBox.warning(self, 'Template Failed',
                                f'Failed to save template:\n{e}')

    @Slot()
    def _import_library(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, 'Import Tag Library', '',
            'Tag Library Files (*.xlsx *.csv);;All Files (*)')
        if not file_path:
            return
        if file_path.lower().endswith('.xlsx'):
            parsed = self._parse_xlsx(file_path)
        else:
            parsed = self._parse_csv(file_path)
        if parsed is None:
            return

        tags, categories, category_by_tag, aliases, implications, profiles = parsed

        choice_dialog = QMessageBox(self)
        choice_dialog.setWindowTitle('Import Tag Library')
        choice_dialog.setText(
            f'Found {len(tags)} tag(s), {len(categories)} '
            f'category/categories, {len(aliases)} alias(es), and '
            f'{len(implications)} implication rule(s), and '
            f'{len(profiles)} profile(s).\n\n'
            'Replace All — replaces your entire Tag Library with the imported data.\n'
            'Merge — adds new tags, categories, aliases, implications, and profiles '
            'without removing existing ones.')
        replace_button = choice_dialog.addButton(
            'Replace All', QMessageBox.ButtonRole.DestructiveRole)
        merge_button = choice_dialog.addButton(
            'Merge', QMessageBox.ButtonRole.AcceptRole)
        choice_dialog.addButton(QMessageBox.StandardButton.Cancel)
        choice_dialog.exec()

        clicked = choice_dialog.clickedButton()
        if clicked is replace_button:
            self.tag_library_model.load_all_data(
                tags, categories, category_by_tag,
                aliases=aliases, implications=implications, profiles=profiles)
            QMessageBox.information(
                self, 'Import Successful',
                f'Replaced Tag Library with '
                f'{len(self.tag_library_model.tags)} tag(s), '
                f'{len(self.tag_library_model.categories)} category/categories, '
                f'{len(self.tag_library_model.aliases)} alias(es), and '
                f'{len(self.tag_library_model.implications)} implication rule(s), '
                f'and {len(self.tag_library_model.profiles)} profile(s).')
        elif clicked is merge_button:
            self._merge_import(tags, categories, category_by_tag, aliases,
                               implications, profiles)

    def _merge_import(self, tags: list, categories: list,
                      category_by_tag: dict, aliases: dict[str, str] | None = None,
                      implications: dict[str, list[str]] | None = None,
                      profiles: dict[str, dict[str, str]] | None = None):
        """Add new tags/categories from import without removing existing data."""
        existing_tags = set(self.tag_library_model.tags)
        existing_cat_names = {
            c['name'].casefold(): c['id']
            for c in self.tag_library_model.categories
        }
        added_categories = 0

        # Map imported category IDs → existing/new IDs (matched by name)
        cat_id_map: dict[str, str] = {}
        for cat in categories:
            name = str(cat.get('name', '')).strip()
            color = str(cat.get('color', ''))
            old_id = str(cat.get('id', ''))
            if not name:
                continue
            if name.casefold() in existing_cat_names:
                cat_id_map[old_id] = existing_cat_names[name.casefold()]
            else:
                self.tag_library_model.add_category(name, color)
                added_categories += 1
                # add_category re-builds categories; find the new ID by name
                for c in self.tag_library_model.categories:
                    if c['name'].casefold() == name.casefold():
                        cat_id_map[old_id] = c['id']
                        break

        # Add new tags
        new_tags = [t for t in tags if t not in existing_tags]
        if new_tags:
            self.tag_library_model.add_tags(new_tags)

        # Assign categories for newly imported tags
        tags_by_new_cat: dict[str, list[str]] = {}
        for tag, old_cat_id in category_by_tag.items():
            new_cat_id = cat_id_map.get(old_cat_id)
            if new_cat_id and tag in new_tags:
                tags_by_new_cat.setdefault(new_cat_id, []).append(tag)
        for cat_id, tag_list in tags_by_new_cat.items():
            self.tag_library_model.assign_category(tag_list, cat_id)

        added_aliases = 0
        if aliases:
            existing_aliases = self.tag_library_model.aliases
            for alias, canonical in aliases.items():
                if alias not in existing_aliases:
                    if self.tag_library_model.add_alias(alias, canonical):
                        added_aliases += 1

        added_implications = 0
        if implications:
            for tag, implied_tags in implications.items():
                if self.tag_library_model.add_implication(tag, implied_tags):
                    added_implications += 1

        added_profiles = 0
        added_profile_mappings = 0
        if profiles:
            existing_profiles = self.tag_library_model.get_profiles()
            for profile_name, mapping in profiles.items():
                if profile_name not in existing_profiles:
                    if self.tag_library_model.add_profile(profile_name):
                        added_profiles += 1
                    existing_profiles[profile_name] = {}
                existing_mapping = existing_profiles.get(profile_name, {})
                for original, replacement in mapping.items():
                    if original in existing_mapping:
                        continue
                    if self.tag_library_model.set_profile_mapping(
                            profile_name, original, replacement):
                        added_profile_mappings += 1
                        existing_mapping[original] = replacement

        added_count = len(new_tags)
        QMessageBox.information(
            self, 'Merge Successful',
            f'Added {added_count} new tag(s), {added_categories} '
            f'category/categories, {added_aliases} alias(es), and '
            f'{added_implications} implication rule(s), {added_profiles} '
            f'profile(s), and {added_profile_mappings} profile mapping(s) '
            'to your Tag Library.')

    def _parse_xlsx(self, file_path: str):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.warning(self, 'Import Failed', f'Could not read file:\n{e}')
            return None

        tags = []
        tag_to_cat_name = {}
        if 'Tags' in wb.sheetnames:
            ws = wb['Tags']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else '' for c in rows[0]]
                tag_col = header.index('tag') if 'tag' in header else None
                cat_col = header.index('category') if 'category' in header else None
                for row in rows[1:]:
                    tag = (str(row[tag_col]).strip()
                           if tag_col is not None and row[tag_col] else '')
                    if not tag or tag == 'None':
                        continue
                    tags.append(tag)
                    if cat_col is not None and row[cat_col]:
                        cat_name = str(row[cat_col]).strip()
                        if cat_name and cat_name != 'None':
                            tag_to_cat_name[tag] = cat_name

        cat_name_to_color = {}
        if 'Categories' in wb.sheetnames:
            ws = wb['Categories']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else '' for c in rows[0]]
                name_col = header.index('name') if 'name' in header else None
                color_col = next((i for i, h in enumerate(header)
                                  if h in ('color', 'color (hex)')), None)
                for row in rows[1:]:
                    name = (str(row[name_col]).strip()
                            if name_col is not None and row[name_col] else '')
                    if not name or name == 'None':
                        continue
                    color = ''
                    if color_col is not None and row[color_col]:
                        color = str(row[color_col]).strip()
                        if color == 'None':
                            color = ''
                    cat_name_to_color[name] = color

        cat_name_to_id = {}
        categories = []
        for name, color in cat_name_to_color.items():
            new_id = str(uuid4())
            cat_name_to_id[name] = new_id
            categories.append({'id': new_id, 'name': name, 'color': color})

        for tag, cat_name in tag_to_cat_name.items():
            if cat_name not in cat_name_to_id:
                new_id = str(uuid4())
                cat_name_to_id[cat_name] = new_id
                categories.append({'id': new_id, 'name': cat_name, 'color': ''})

        category_by_tag = {
            tag: cat_name_to_id[cat_name]
            for tag, cat_name in tag_to_cat_name.items()
            if cat_name in cat_name_to_id
        }

        aliases = {}
        if 'Aliases' in wb.sheetnames:
            ws = wb['Aliases']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else '' for c in rows[0]]
                alias_col = header.index('alias') if 'alias' in header else None
                canonical_col = (header.index('canonical')
                                 if 'canonical' in header else None)
                for row in rows[1:]:
                    alias = (str(row[alias_col]).strip()
                             if alias_col is not None and row[alias_col] else '')
                    canonical = (str(row[canonical_col]).strip()
                                 if canonical_col is not None
                                 and row[canonical_col] else '')
                    if (alias and canonical and alias != 'None'
                            and canonical != 'None'):
                        aliases[alias] = canonical

        implications = {}
        if 'Implications' in wb.sheetnames:
            ws = wb['Implications']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else '' for c in rows[0]]
                tag_col = header.index('tag') if 'tag' in header else None
                implies_col = header.index('implies') if 'implies' in header else None
                for row in rows[1:]:
                    tag = (str(row[tag_col]).strip()
                           if tag_col is not None and row[tag_col] else '')
                    implies_str = (str(row[implies_col]).strip()
                                   if implies_col is not None and row[implies_col]
                                   else '')
                    if (tag and implies_str and tag != 'None'
                            and implies_str != 'None'):
                        implied = [t.strip() for t in implies_str.split(',')
                                   if t.strip()]
                        if implied:
                            implications[tag] = implied

        profiles = {}
        if 'Profiles' in wb.sheetnames:
            ws = wb['Profiles']
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(c).strip().lower() if c else '' for c in rows[0]]
                profile_col = header.index('profile') if 'profile' in header else None
                original_col = header.index('original') if 'original' in header else None
                replacement_col = (header.index('replacement')
                                   if 'replacement' in header else None)
                for row in rows[1:]:
                    profile = (str(row[profile_col]).strip()
                               if profile_col is not None and row[profile_col] else '')
                    original = (str(row[original_col]).strip()
                                if original_col is not None and row[original_col] else '')
                    replacement = (str(row[replacement_col]).strip()
                                   if replacement_col is not None
                                   and row[replacement_col] else '')
                    if (profile and original and profile != 'None'
                            and original != 'None'):
                        if replacement == 'None':
                            replacement = ''
                        if profile not in profiles:
                            profiles[profile] = {}
                        profiles[profile][original] = replacement

        wb.close()
        return tags, categories, category_by_tag, aliases, implications, profiles

    def _parse_csv(self, file_path: str):
        try:
            with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception as e:
            QMessageBox.warning(self, 'Import Failed',
                                f'Could not read file:\n{e}')
            return None

        if not rows or 'tag' not in (rows[0] if rows else {}):
            QMessageBox.warning(
                self, 'Import Failed',
                'The selected file is not a valid Tag Library CSV.\n'
                'Expected a header row with at least a "tag" column.')
            return None

        tags: list[str] = []
        cat_name_to_color: dict[str, str] = {}
        tag_to_cat_name: dict[str, str] = {}
        for row in rows:
            tag = row.get('tag', '').strip()
            if not tag:
                continue
            tags.append(tag)
            cat_name = row.get('category', '').strip()
            cat_color = row.get('color', '').strip()
            if cat_name:
                cat_name_to_color.setdefault(cat_name, cat_color)
                tag_to_cat_name[tag] = cat_name

        cat_name_to_id: dict[str, str] = {}
        categories: list[dict] = []
        for name, color in cat_name_to_color.items():
            new_id = str(uuid4())
            cat_name_to_id[name] = new_id
            categories.append({'id': new_id, 'name': name, 'color': color})

        category_by_tag = {
            tag: cat_name_to_id[cat_name]
            for tag, cat_name in tag_to_cat_name.items()
            if cat_name in cat_name_to_id
        }
        return tags, categories, category_by_tag, {}, {}, {}

    # ── Categories tab slots ──────────────────────────────────────────────

    @Slot()
    def _refresh_categories_tab(self):
        previous_id = None
        current_item = self.categories_list.currentItem()
        if current_item is not None:
            previous_id = current_item.data(Qt.ItemDataRole.UserRole)

        categories = self.tag_library_model.get_categories()
        self.categories_list.clear()
        for cat in categories:
            item = QListWidgetItem(cat['name'])
            item.setData(Qt.ItemDataRole.UserRole, cat['id'])
            color = QColor(cat['color'])
            if color.isValid():
                item.setForeground(color)
            self.categories_list.addItem(item)

        self.category_count_label.setText(f'{len(categories)} Categories')

        if not previous_id:
            return
        for i in range(self.categories_list.count()):
            item = self.categories_list.item(i)
            if item.data(Qt.ItemDataRole.UserRole) == previous_id:
                self.categories_list.setCurrentItem(item)
                return

    @Slot()
    def _add_alias(self):
        aliases_text = self.new_alias_input.text().strip()
        canonical = self.new_canonical_input.text().strip()
        if not aliases_text or not canonical:
            return
        aliases = [a.strip() for a in aliases_text.split(',') if a.strip()]
        if not aliases:
            return
        for alias in aliases:
            if alias == canonical:
                QMessageBox.warning(
                    self, 'Invalid Alias',
                    f'An alias cannot be the same as its canonical tag ("{canonical}").')
                return
        for alias in aliases:
            self.tag_library_model.add_alias(alias, canonical)
        self.new_alias_input.clear()
        self.new_canonical_input.clear()

    @Slot()
    def _filter_aliases_table(self):
        text = self.alias_search_input.text().lower()
        for row in range(self.aliases_table.rowCount()):
            item0 = self.aliases_table.item(row, 0)
            item1 = self.aliases_table.item(row, 1)
            t0 = item0.text().lower() if item0 else ''
            t1 = item1.text().lower() if item1 else ''
            self.aliases_table.setRowHidden(row, text not in t0 and text not in t1)

    @Slot()
    def _edit_selected_alias(self):
        selected_rows = sorted(set(idx.row()
                                   for idx in self.aliases_table.selectedIndexes()))
        if len(selected_rows) != 1:
            return
        row = selected_rows[0]
        alias_item = self.aliases_table.item(row, 0)
        canonical_item = self.aliases_table.item(row, 1)
        if not alias_item or not canonical_item:
            return
        self.new_alias_input.setText(alias_item.text())
        self.new_canonical_input.setText(canonical_item.text())
        self.tag_library_model.remove_aliases([alias_item.text()])
        self.new_alias_input.setFocus()

    @Slot()
    def _remove_selected_aliases(self):
        selected_rows = set(index.row()
                            for index in self.aliases_table.selectedIndexes())
        if not selected_rows:
            return
        aliases_to_remove = []
        for row in selected_rows:
            alias_item = self.aliases_table.item(row, 0)
            if alias_item:
                aliases_to_remove.append(alias_item.text())
        if aliases_to_remove:
            self.tag_library_model.remove_aliases(aliases_to_remove)

    @Slot()
    def _refresh_aliases_tab(self):
        self.aliases_table.setSortingEnabled(False)
        self.aliases_table.setRowCount(0)
        aliases = self.tag_library_model.get_aliases()
        for alias, canonical in sorted(aliases.items()):
            row = self.aliases_table.rowCount()
            self.aliases_table.insertRow(row)
            alias_item = QTableWidgetItem(alias)
            alias_item.setFlags(alias_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            canonical_item = QTableWidgetItem(canonical)
            canonical_item.setFlags(
                canonical_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.aliases_table.setItem(row, 0, alias_item)
            self.aliases_table.setItem(row, 1, canonical_item)
        self.aliases_table.setSortingEnabled(True)
        self.alias_count_label.setText(f'{len(aliases)} Aliases')
        self._filter_aliases_table()

    @Slot()
    def _add_implication(self):
        tag_text = self.implication_tag_input.text().strip()
        implied_text = self.implication_implied_input.text().strip()
        if not tag_text or not implied_text:
            return
        trigger_tags = [t.strip() for t in tag_text.split(',') if t.strip()]
        implied_tags = [t.strip() for t in implied_text.split(',') if t.strip()]
        if not trigger_tags or not implied_tags:
            return
        if any('*' in t or '?' in t for t in implied_tags):
            QMessageBox.warning(self, 'Invalid Implication',
                                'Implied tags cannot contain wildcards '
                                '(* or ?). Wildcards are only allowed on the '
                                'trigger (left) side.')
            return
        added_any = False
        skipped_self = []
        for trigger in trigger_tags:
            remaining = [t for t in implied_tags if t != trigger]
            if not remaining:
                skipped_self.append(trigger)
                continue
            if self.tag_library_model.add_implication(trigger, remaining):
                added_any = True
        if skipped_self and not added_any:
            QMessageBox.warning(self, 'Invalid Implication',
                                'A tag cannot imply itself.')
            return
        self.implication_tag_input.clear()
        self.implication_implied_input.clear()

    @Slot()
    def _filter_implications_table(self):
        text = self.implication_search_input.text().lower()
        for row in range(self.implications_table.rowCount()):
            item0 = self.implications_table.item(row, 0)
            item1 = self.implications_table.item(row, 1)
            t0 = item0.text().lower() if item0 else ''
            t1 = item1.text().lower() if item1 else ''
            self.implications_table.setRowHidden(
                row, text not in t0 and text not in t1)

    @Slot()
    def _edit_selected_implication(self):
        selected_rows = sorted(set(
            idx.row() for idx in self.implications_table.selectedIndexes()))
        if len(selected_rows) != 1:
            return
        row = selected_rows[0]
        tag_item = self.implications_table.item(row, 0)
        implies_item = self.implications_table.item(row, 1)
        if not tag_item or not implies_item:
            return
        self.implication_tag_input.setText(tag_item.text())
        self.implication_implied_input.setText(implies_item.text())
        self.tag_library_model.remove_implication_rule(tag_item.text())
        self.implication_tag_input.setFocus()

    @Slot()
    def _remove_selected_implication_rules(self):
        selected_rows = set(index.row()
                            for index in self.implications_table.selectedIndexes())
        if not selected_rows:
            return
        tags_to_remove = []
        for row in selected_rows:
            item = self.implications_table.item(row, 0)
            if item:
                tags_to_remove.append(item.text())
        for tag in tags_to_remove:
            self.tag_library_model.remove_implication_rule(tag)

    @Slot()
    def _refresh_implications_tab(self):
        self.implications_table.setSortingEnabled(False)
        self.implications_table.setRowCount(0)
        implications = self.tag_library_model.get_implications()
        for tag, implied_tags in sorted(implications.items()):
            row = self.implications_table.rowCount()
            self.implications_table.insertRow(row)
            self.implications_table.setItem(row, 0, QTableWidgetItem(tag))
            self.implications_table.setItem(
                row, 1, QTableWidgetItem(', '.join(implied_tags)))
        self.implications_table.setSortingEnabled(True)
        count = len(implications)
        self.implication_count_label.setText(
            f'{count} Implication Rule{"s" if count != 1 else ""}')
        self._filter_implications_table()

    @Slot()
    def _request_add_category(self):
        dialog = CategoryDialog(self, title='Add Category')
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        name, color = dialog.get_values()
        if name:
            self.tag_library_model.add_category(name, color)

    @Slot()
    def _request_edit_category(self):
        current_item = self.categories_list.currentItem()
        if current_item is None:
            return
        category_id = current_item.data(Qt.ItemDataRole.UserRole)
        for cat in self.tag_library_model.get_categories():
            if cat['id'] != category_id:
                continue
            dialog = CategoryDialog(self, title='Edit Category',
                                    name=cat['name'],
                                    color=cat['color'] or '#ffffff')
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
            name, color = dialog.get_values()
            if name:
                self.tag_library_model.edit_category(category_id, name, color)
            return

    @Slot()
    def _request_remove_category(self):
        current_item = self.categories_list.currentItem()
        if current_item is None:
            return
        category_id = current_item.data(Qt.ItemDataRole.UserRole)
        category_name = current_item.text()
        reply = get_confirmation_dialog_reply(
            title='Remove Category',
            question=f'Remove category "{category_name}" from the Tag Library? '
                     'Assigned tags will keep the tag and lose the category.')
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.tag_library_model.remove_category(category_id)

    @Slot()
    def _persist_category_order(self, *args):
        ordered_ids = [
            self.categories_list.item(i).data(Qt.ItemDataRole.UserRole)
            for i in range(self.categories_list.count())
        ]
        self.tag_library_model.set_category_order(ordered_ids)
